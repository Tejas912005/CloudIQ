"""
backend/routers/exports.py
--------------------------
Export endpoints: PDF dashboard report and Excel workbook.
Both are now fully implemented using reportlab (PDF) and openpyxl (Excel).
"""

from __future__ import annotations
import io
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session
from core.auth import verify_api_key
from core.database import get_db
from models.models import CloudResource, CostHistory, Recommendation, AnomalyRecord

router = APIRouter(prefix="/api", tags=["Exports"])


@router.post("/export/pdf", dependencies=[Depends(verify_api_key)])
def export_pdf(db: Session = Depends(get_db)):
    """
    Generate a PDF dashboard report containing:
    - Executive summary (resource counts, total cost, health breakdown)
    - Top 10 resources by monthly cost
    - Active recommendations
    - Recent cost anomalies
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer,
            Table, TableStyle, HRFlowable,
        )

        # â”€â”€ Fetch data â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        resources    = db.query(CloudResource).all()
        recs         = db.query(Recommendation).order_by(Recommendation.priority).limit(10).all()
        anomalies    = db.query(AnomalyRecord).limit(10).all()
        cost_history = db.query(CostHistory).order_by(CostHistory.date.desc()).limit(7).all()

        total_cost  = round(sum(r.monthly_cost or 0 for r in resources), 2)
        idle_count  = sum(1 for r in resources if r.status == "Idle")
        over_count  = sum(1 for r in resources if r.status == "Over-Utilized")
        healthy     = len(resources) - idle_count - over_count
        top_cost    = sorted(resources, key=lambda r: r.monthly_cost or 0, reverse=True)[:10]

        # â”€â”€ Document setup â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        buf  = io.BytesIO()
        doc  = SimpleDocTemplate(buf, pagesize=A4,
                                 leftMargin=2*cm, rightMargin=2*cm,
                                 topMargin=2*cm, bottomMargin=2*cm)
        styles = getSampleStyleSheet()
        story  = []

        # Header
        H1 = ParagraphStyle("H1", parent=styles["Title"],
                             fontSize=22, textColor=colors.HexColor("#63b2ff"),
                             spaceAfter=4)
        H2 = ParagraphStyle("H2", parent=styles["Heading2"],
                             fontSize=13, textColor=colors.HexColor("#63b2ff"),
                             spaceBefore=12, spaceAfter=4)
        BODY = ParagraphStyle("BODY", parent=styles["Normal"],
                              fontSize=9, textColor=colors.HexColor("#94a3b8"))
        generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        story.append(Paragraph("â˜  CloudIQ Dashboard Report", H1))
        story.append(Paragraph(f"Generated: {generated}", BODY))
        story.append(HRFlowable(width="100%", thickness=1,
                                color=colors.HexColor("#1e3a5f"), spaceAfter=12))

        # Executive Summary
        story.append(Paragraph("Executive Summary", H2))
        summary_data = [
            ["Metric", "Value"],
            ["Total Resources",       str(len(resources))],
            ["Healthy",               str(healthy)],
            ["Idle (waste candidate)", str(idle_count)],
            ["Over-Utilized",         str(over_count)],
            ["Total Monthly Cost",    f"${total_cost:,.2f}"],
        ]
        story.append(_make_table(summary_data))
        story.append(Spacer(1, 0.4*cm))

        # Top 10 Resources by Cost
        story.append(Paragraph("Top 10 Resources by Monthly Cost", H2))
        res_data = [["Name", "Type", "Region", "Status", "Monthly Cost"]]
        for r in top_cost:
            res_data.append([
                r.name, r.resource_type, r.region,
                r.status, f"${r.monthly_cost:,.2f}"
            ])
        story.append(_make_table(res_data))
        story.append(Spacer(1, 0.4*cm))

        # Recommendations
        if recs:
            story.append(Paragraph("Active Recommendations", H2))
            rec_data = [["Resource", "Action", "Priority", "Est. Savings"]]
            for r in recs:
                savings = f"${r.estimated_savings:,.2f}" if r.estimated_savings else "â€”"
                rec_data.append([r.resource_name, r.action[:60], r.priority, savings])
            story.append(_make_table(rec_data))
            story.append(Spacer(1, 0.4*cm))

        # Recent Cost History
        if cost_history:
            story.append(Paragraph("Recent Daily Cost (Last 7 Days)", H2))
            hist_data = [["Date", "Daily Cost", "Anomaly"]]
            for ch in reversed(cost_history):
                hist_data.append([
                    ch.date,
                    f"${ch.daily_cost:,.2f}",
                    "âš  Yes" if ch.is_anomaly else "âœ“ No",
                ])
            story.append(_make_table(hist_data))

        doc.build(story)
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=cloudiq_report.pdf"},
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {str(e)}")


def _make_table(data: list) -> "Table":
    """Helper: styled table for PDF output."""
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors

    t = Table(data, hAlign="LEFT")
    style = TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#0d2137")),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.HexColor("#63b2ff")),
        ("FONTSIZE",    (0, 0), (-1, 0),  9),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("TEXTCOLOR",   (0, 1), (-1, -1), colors.HexColor("#94a3b8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#0a1628"), colors.HexColor("#0d1f38")]),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#1e3a5f")),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ])
    t.setStyle(style)
    return t


@router.post("/export/xlsx", dependencies=[Depends(verify_api_key)])
def export_xlsx(db: Session = Depends(get_db)):
    """
    Generate an Excel workbook with four sheets:
    - Resources (all columns)
    - Cost History (90-day series)
    - Recommendations
    - Anomalies
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        resources    = db.query(CloudResource).all()
        cost_history = db.query(CostHistory).order_by(CostHistory.date).all()
        recs         = db.query(Recommendation).all()
        anomalies    = db.query(AnomalyRecord).all()

        wb = openpyxl.Workbook()

        # â”€â”€ Styling helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        HDR_FILL  = PatternFill("solid", fgColor="0D2137")
        HDR_FONT  = Font(bold=True, color="63B2FF", size=10)
        EVEN_FILL = PatternFill("solid", fgColor="0A1628")
        ODD_FILL  = PatternFill("solid", fgColor="0D1F38")
        DATA_FONT = Font(color="94A3B8", size=9)
        CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
        THIN_BORDER = Border(
            left=Side(style="thin", color="1E3A5F"),
            right=Side(style="thin", color="1E3A5F"),
            top=Side(style="thin", color="1E3A5F"),
            bottom=Side(style="thin", color="1E3A5F"),
        )

        def write_sheet(ws, headers, rows):
            ws.append(headers)
            for col_num, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill   = HDR_FILL
                cell.font   = HDR_FONT
                cell.alignment = CENTER
                cell.border = THIN_BORDER
                ws.column_dimensions[get_column_letter(col_num)].width = 18
            for r_idx, row in enumerate(rows, 2):
                ws.append(row)
                fill = EVEN_FILL if r_idx % 2 == 0 else ODD_FILL
                for c_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.fill   = fill
                    cell.font   = DATA_FONT
                    cell.border = THIN_BORDER

        # Sheet 1: Resources
        ws1 = wb.active
        ws1.title = "Resources"
        write_sheet(ws1,
            ["Name", "UID", "Type", "Provider", "Region", "Status",
             "CPU%", "Mem%", "Monthly Cost ($)", "Risk Score", "Public"],
            [[r.name, r.resource_uid, r.resource_type, r.provider, r.region,
              r.status, round(r.cpu_usage or 0, 1), round(r.memory_usage or 0, 1),
              round(r.monthly_cost or 0, 2), round(r.risk_score or 0, 2),
              "Yes" if r.public_access else "No"] for r in resources]
        )

        # Sheet 2: Cost History
        ws2 = wb.create_sheet("Cost History")
        write_sheet(ws2,
            ["Date", "Daily Cost ($)", "Anomaly"],
            [[ch.date, round(ch.daily_cost, 2),
              "Yes" if ch.is_anomaly else "No"] for ch in cost_history]
        )

        # Sheet 3: Recommendations
        ws3 = wb.create_sheet("Recommendations")
        write_sheet(ws3,
            ["Resource", "Action", "Priority", "Category", "Est. Savings ($)", "Reason"],
            [[r.resource_name, r.action, r.priority, r.category or "cost",
              round(r.estimated_savings or 0, 2), r.reason] for r in recs]
        )

        # Sheet 4: Anomalies
        if anomalies:
            ws4 = wb.create_sheet("Anomalies")
            write_sheet(ws4,
                ["Resource ID", "Type", "Z-Score", "Deviation", "Severity", "Description", "Detected At"],
                [[a.resource_id, a.anomaly_type,
                  round(a.z_score or 0.0, 2),
                  round(a.deviation or 0.0, 2),
                  a.severity or "medium",
                  a.description or "",
                  str(a.created_at or "")
                  ] for a in anomalies]
            )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=cloudiq_export.xlsx"},
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")
